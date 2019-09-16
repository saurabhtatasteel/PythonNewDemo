from openpyxl import *
from tkinter import *
class MyGUI:

    wb = load_workbook('C:\Trials\Python Demos\TKinter\RF.xlsx')

    def __init__(self,myWindow):
        MyGUI.sheet = MyGUI.wb.active 

    def excel(self): 
        MyGUI.sheet.column_dimensions['A'].width = 30
        MyGUI.sheet.column_dimensions['B'].width = 10
        MyGUI.sheet.column_dimensions['C'].width = 10
        MyGUI.sheet.column_dimensions['D'].width = 20
        MyGUI.sheet.column_dimensions['E'].width = 20
        MyGUI.sheet.column_dimensions['F'].width = 40
        MyGUI.sheet.column_dimensions['G'].width = 50


        MyGUI.sheet.cell(row=1, column=1).value = "Name"
        MyGUI.sheet.cell(row=1, column=2).value = "Course"
        MyGUI.sheet.cell(row=1, column=3).value = "Semester"
        MyGUI.sheet.cell(row=1, column=4).value = "Form Number"
        MyGUI.sheet.cell(row=1, column=5).value = "Contact Number"
        MyGUI.sheet.cell(row=1, column=6).value = "Email id"
        MyGUI.sheet.cell(row=1, column=7).value = "Address"


myWindow = Tk()
myWindow.geometry("472x600")
myWindow.resizable(0,0)
mygui = MyGUI(myWindow)
myWindow.title("Registration Form")
myWindow.configure(bg = "silver")



heading = Label(myWindow, text="Form") 
name = Label(myWindow, text="Name") 
course = Label(myWindow, text="Course")
sem = Label(myWindow, text="Semester")
form_no = Label(myWindow, text="Form No.") 
contact_no = Label(myWindow, text="Contact No.")
email_id = Label(myWindow, text="Email id") 
address = Label(myWindow, text="Address") 

heading.grid(row=0, column=1) 
name.grid(row=1, column=0) 
course.grid(row=2, column=0) 
sem.grid(row=3, column=0) 
form_no.grid(row=4, column=0) 
contact_no.grid(row=5, column=0) 
email_id.grid(row=6, column=0) 
address.grid(row=7, column=0)

name = Entry(myWindow)
name.grid(row = 1,column = 1)
course = Entry(myWindow)
course.grid(row = 2,column = 1)
sem = Entry(myWindow)
sem.grid(row = 3,column = 1)
form_no = Entry(myWindow)
form_no.grid(row = 4,column = 1)
contact_no = Entry(myWindow)
contact_no.grid(row = 5,column = 1)
email_id = Entry(myWindow)
email_id.grid(row = 6,column = 1)
address = Entry(myWindow)
address.grid(row = 7,column = 1)

submitButton =Button(myWindow ,text="Submit")
submitButton.grid(row=9,column=1)

mygui.excel() 

#Entering the event main loop
myWindow.mainloop()


